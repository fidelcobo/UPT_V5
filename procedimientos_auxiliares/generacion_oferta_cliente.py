# Este fichero contiene sólo procedimientos para hacer la oferta de cliente. Nuevo en V4
import openpyxl
from typing import List
from aux_class import ElementosOferta, DatosGenerales, FullElementosOferta
import os
from os.path import dirname
from datetime import datetime, timedelta


def generar_oferta_cliente(datos_clases: List[FullElementosOferta], datos_generales, instance):
    """
    Esta rutina nos genera un oferta Excel entregable al cliente. Para ello se parte de la listas de los
    de los ítems de la quote. Se consolidan aquellos artículos en los que se
    produce coincidencia de fabricante, código y fechas de principio y fin del servicio
    :param datos_clases: Lista de entradas consolidadas de la oferta
    :param datos_generales: Nombre de la oferta, del cliente, BID y AM
    :param instance: Clase Qt5 de origen. Usado para enviar mensajes de pantalla al usuario
    :return: Fichero oferta_cliente.xlsx de oferta
    """

    lista_articulos_oferta = []

    for element in datos_clases:
        pvp_total = element.qty * element.total_sell_price
        item = ElementosOferta(element.in_csv, element.manufacturer, element.code, element.qty, element.init_date,
                               element.end_date, element.uptime, pvp_total)
        lista_articulos_oferta.append(item)

    lista_articulos_oferta.sort(key=lambda x: x.manufacturer)  # Ordenación alfabética por fabricante
    book = hacer_libro_de_oferta_de_cliente(lista_articulos_oferta, datos_generales, instance)
    if book:
        book.close()
    return book


def hacer_libro_de_oferta_de_cliente(lista_elementos, datos_generales: DatosGenerales, instance):
    """
    Éste es un procedimiento auxiliar de generar_oferta_cliente. Realiza la composición del book Excel de la
    oferta de cliente partiendo de los datos de la misma. Asimismo, ajusta la longitud de filas Excel de la
    oferta para que quede presentable y escribe la fecha actual de composición.

    :param lista_elementos: Ítems de la oferta
    :param datos_generales: Nombre del cliente, proyecto, BID, AM
    :param instance: Usado para enviar pantalla de error en caso de una oferta demasiado grande
    :return: libro Excel con oferta de cliente (openpyxl.Workbook)
    """

    # Comenzamos definiendo los parámetros de filas y columnas de la plantilla de oferta
    FIRST_ROW = 22
    fabricante = 'B'
    equipo = 'F'
    unid = 'N'
    servicio = 'P'
    fecha_inicio = 'AC'
    fecha_fin = 'AH'
    importe = 'AM'
    columna_precio_total_oferta = 'AM'

    quote_name = 'L6'
    cliente = 'L7'
    bid = 'L8'
    version = 'R9'
    am = 'T11'
    propuesta = 'B18'
    duracion_oferta = 'AL13'
    fecha_limite_validez = 'AL14'
    columna_fecha = 'AM'
    columna_guia = 'AH'
    base_dir = dirname(os.path.abspath(os.path.dirname(__file__)))
    template_file = os.path.join(base_dir + '/template_oferta_cliente.xlsx')

    total = 0

    # Abrimos el libro
    try:
        book = openpyxl.load_workbook(template_file, data_only=True)
        sheet = book.get_sheet_by_name('OFERTA CLIENTE')

        # Ahora buscamos el máximo número de ítems de la plantilla
        max_items = 0
        for i in range(1000):  # Como mucho, 1000
            row = str(FIRST_ROW + i)
            celda = columna_guia + str(row)
            if sheet[celda].value == 'Total Venta (EUR)':
                max_items = i - 4
                break
        if max_items == 0:  # El template de formato de oferta falla
            print('Template de formato de oferta incorrecto')
            return None

        # Verificamos si el número máximo de ítems es compatible con la oferta
        if max_items < len(lista_elementos):
            instance.signals.error_fichero.emit("Demasiados ítems. \nNo se ha generado oferta de cliente")
            return None

        # Como está en orden, rellenamos con los datos de las lista recibidas
        for i in range(len(lista_elementos)):
            row = str(FIRST_ROW + i)
            sheet[fabricante + row] = lista_elementos[i].manufacturer  # Escribimos el fabricante
            sheet[equipo + row] = lista_elementos[i].code  # Código de equipo
            sheet[unid + row] = lista_elementos[i].qty  # Cantidad
            sheet[servicio + row] = lista_elementos[i].uptime  # Servicio Uptime
            sheet[fecha_inicio + row] = lista_elementos[i].init_date  # Fecha de inicio del servicio
            sheet[fecha_fin + row] = lista_elementos[i].end_date  # Fecha de final
            sheet[importe + row] = lista_elementos[i].total_price  # Precio total
            total += lista_elementos[i].total_price

        # Marcamos ahora el total de la oferta
        fila_precio_total = max_items + FIRST_ROW + 4
        celda_precio_total = str(columna_precio_total_oferta) + str(fila_precio_total)
        sheet[celda_precio_total] = total

        # Y ahora las fechas
        fila_fecha = max_items + FIRST_ROW + 9
        celda_fecha = str(columna_fecha) + str(fila_fecha)
        sheet[celda_fecha].number_format = 'dd/mm/yyyy'  # Fijamos el formato de la celda de fecha
        sheet[celda_fecha] = datetime.today()  # La fecha de ejecución del programa
        validez_oferta = timedelta(days=sheet[duracion_oferta].value)
        sheet[fecha_limite_validez] = datetime.today() + validez_oferta  # La fecha máxima de validez de la oferta

        # Ya hemos rellenado los datos de los ítems de la oferta. Ahora hay que arreglar el formato
        # Se trata de borrar las filas sobrantes para que el formato quede bien.
        num_filas_a_borrar = max_items - len(lista_elementos)
        if num_filas_a_borrar > (max_items - 25):  # Dejamos al menos 25 líneas si la oferta es pequeña
            num_filas_a_borrar = max_items - 25
        fila_comienzo = FIRST_ROW + len(lista_elementos) + 1
        sheet.delete_rows(fila_comienzo, num_filas_a_borrar)

        # Ahora ponemos los datos generales de la oferta
        sheet[quote_name] = datos_generales.nombre_oferta
        sheet[cliente] = datos_generales.cliente
        sheet[bid] = datos_generales.bid
        sheet[version] = datos_generales.version
        sheet[am] = datos_generales.am.value
        sheet[propuesta] = datos_generales.nombre_oferta

        # Finalmente, borramos las líneas sobrantes al final de la oferta para que quede más limpia
        print(sheet.max_row)
        # ultima_fila = sheet.max_row
        # fila_comienzo = fila_fecha + 3
        # num_filas_a_borrar = fila_comienzo - ultima_fila
        # sheet.delete_rows(ultima_fila + 1, num_filas_a_borrar)

        # Sólo queda cerrar el libro y devolverlo
        book.close()
        return book

    except Exception as e:
        print('Excepción:' + e.__repr__())
        return None


def obtener_datos_generales(hoja_resumen):
    quote_name = hoja_resumen['C6'].value
    customer = hoja_resumen['C7'].value
    bid = hoja_resumen['C8'].value
    version = hoja_resumen['E9'].value
    am = hoja_resumen['E11']

    datos = DatosGenerales(nombre_oferta=quote_name, cliente=customer, bid=bid, version=version, am=am)
    return datos
